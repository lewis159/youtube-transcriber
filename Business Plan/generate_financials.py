"""
YT Transcriber — Financial Model Generator
All data cells use Excel formulas referencing the Assumptions sheet.
Change any yellow input cell and the entire workbook updates automatically.

Tiers modelled: Explorer (free/trial), Creator, Studio, Enterprise, Blended Mix
"""
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.chart import BarChart, Reference
from datetime import date

OUTPUT = r"Z:\GIT\Cursor\new project\youtube-transcriber\Business Plan\YT_Transcriber_Financial_Model.xlsx"
TODAY  = date.today().strftime("%d %B %Y")

# ── Brand colours ─────────────────────────────────────────────────────────────
C_DARK   = "1A1A1A"
C_RED    = "E53935"
C_HEAD2  = "2A2A2A"
C_WHITE  = "FFFFFF"
C_LGREY  = "F5F5F5"
C_EGREY  = "E0E0E0"
C_MUTED  = "888888"
C_INPUT  = "FFFDE7"   # yellow  — editable input cell

def fill(h):    return PatternFill("solid", fgColor=h)
def fnt(bold=False, color=C_DARK, size=9, italic=False):
    return Font(name="Calibri", bold=bold, color=color, size=size, italic=italic)
def aln(h="left", wrap=False):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)
def bdr():
    s = Side(border_style="thin", color=C_EGREY)
    return Border(left=s, right=s, top=s, bottom=s)

GBP  = u'"£"#,##0.00'
GBP0 = u'"£"#,##0'
PCT  = '0.0%'
NUM  = '#,##0'

# ── Conditional-formatting styles ─────────────────────────────────────────────
cf_red_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
cf_red_font = Font(name="Calibri", bold=True,  color="C62828", size=9)
cf_grn_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
cf_grn_font = Font(name="Calibri", bold=True,  color="2E7D32", size=9)
cf_amb_fill = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
cf_amb_font = Font(name="Calibri", bold=True,  color="E65100", size=9)

def add_profit_cf(ws, col, r1, r2):
    rng = f"{col}{r1}:{col}{r2}"
    ws.conditional_formatting.add(rng, CellIsRule(operator='lessThan',           formula=['0'], fill=cf_red_fill, font=cf_red_font))
    ws.conditional_formatting.add(rng, CellIsRule(operator='greaterThanOrEqual', formula=['0'], fill=cf_grn_fill, font=cf_grn_font))

def add_margin_cf(ws, col, r1, r2):
    rng = f"{col}{r1}:{col}{r2}"
    ws.conditional_formatting.add(rng, CellIsRule(operator='lessThan',           formula=['0'],     fill=cf_red_fill, font=cf_red_font))
    ws.conditional_formatting.add(rng, CellIsRule(operator='between',            formula=['0','0.2'], fill=cf_amb_fill, font=cf_amb_font))
    ws.conditional_formatting.add(rng, CellIsRule(operator='greaterThanOrEqual', formula=['0.2'],   fill=cf_grn_fill, font=cf_grn_font))

def add_status_cf(ws, col, r1, r2):
    rng = f"{col}{r1}:{col}{r2}"
    top = f"{col}{r1}"
    ws.conditional_formatting.add(rng, FormulaRule(formula=[f'{top}="Loss"'],       fill=cf_red_fill, font=cf_red_font))
    ws.conditional_formatting.add(rng, FormulaRule(formula=[f'{top}="Break-even"'], fill=cf_amb_fill, font=cf_amb_font))
    ws.conditional_formatting.add(rng, FormulaRule(formula=[f'{top}="Growing"'],    fill=cf_grn_fill, font=cf_grn_font))
    ws.conditional_formatting.add(rng, FormulaRule(formula=[f'{top}="Profitable"'], fill=cf_grn_fill, font=cf_grn_font))

# ══════════════════════════════════════════════════════════════════════════════
#  ASSUMPTIONS SHEET ROW MAP
#  Yellow cells = inputs you edit.  Blue cells = auto-calculated formulas.
#
#  Row  │ Content
#  ─────┼──────────────────────────────────────────────────
#    6  │ Creator price
#    7  │ Studio price
#    8  │ Enterprise price
#   11  │ Creator mix %
#   12  │ Studio mix %
#   13  │ Enterprise mix %
#   14  │ Blended ARPU (formula)
#   17  │ Vercel cost
#   18  │ Supabase cost
#   19  │ Domain cost
#   20  │ Total fixed costs (formula)
#   23  │ Explorer variable cost/user
#   24  │ Explorer AI cost/user    (= 0, no AI on trial)
#   25  │ Creator variable cost/user
#   26  │ Creator AI cost/user
#   27  │ Studio variable cost/user
#   28  │ Studio AI cost/user
#   29  │ Enterprise variable cost/user
#   30  │ Enterprise AI cost/user
#   33  │ Stripe % fee
#   34  │ Stripe per-transaction fee
#   37  │ Clerk cost
#   40  │ Blended variable/user (formula)
#   41  │ Blended AI/user (formula)
# ══════════════════════════════════════════════════════════════════════════════

R = {
    "creator_price":  "$B$6",
    "studio_price":   "$B$7",
    "ent_price":      "$B$8",
    "creator_mix":    "$B$11",
    "studio_mix":     "$B$12",
    "ent_mix":        "$B$13",
    "arpu":           "$B$14",
    "fixed":          "$B$20",
    "explorer_var":   "$B$23",
    "explorer_ai":    "$B$24",
    "creator_var":    "$B$25",
    "creator_ai":     "$B$26",
    "studio_var":     "$B$27",
    "studio_ai":      "$B$28",
    "ent_var":        "$B$29",
    "ent_ai":         "$B$30",
    "stripe_pct":     "$B$33",
    "stripe_fxd":     "$B$34",
    "blend_var":      "$B$40",
    "blend_ai":       "$B$41",
}

def A(key):
    """Absolute cross-sheet reference into Assumptions."""
    return f"Assumptions!{R[key]}"

USER_COUNTS = [1, 5, 10, 15, 20, 50, 100, 200, 500, 1000, 2000, 5000]
DATA_START  = 4
DATA_END    = DATA_START + len(USER_COUNTS) - 1

wb = Workbook()
wb.remove(wb.active)


# ══════════════════════════════════════════════════════════════════════════════
#  ASSUMPTIONS SHEET
# ══════════════════════════════════════════════════════════════════════════════
ws_a = wb.create_sheet("Assumptions")
ws_a.sheet_properties.tabColor = C_RED

ws_a.merge_cells("A1:C1")
ws_a["A1"] = "YT Transcriber — Financial Model Assumptions"
ws_a["A1"].fill = fill(C_DARK); ws_a["A1"].font = fnt(bold=True, color=C_WHITE, size=14)
ws_a["A1"].alignment = aln("left"); ws_a.row_dimensions[1].height = 30
ws_a.merge_cells("A2:C2")
ws_a["A2"] = f"Model C: Explorer (free trial) / Creator / Studio / Enterprise     |     {TODAY}"
ws_a["A2"].fill = fill(C_RED); ws_a["A2"].font = fnt(color=C_WHITE, size=9)
ws_a["A2"].alignment = aln("left"); ws_a.row_dimensions[2].height = 16

for ci, h in enumerate(["Assumption", "Value", "Notes"], 1):
    c = ws_a.cell(row=4, column=ci, value=h)
    c.fill = fill(C_HEAD2); c.font = fnt(bold=True, color=C_WHITE, size=9)
    c.alignment = aln("center"); c.border = bdr()
ws_a.row_dimensions[4].height = 20

# (row, label, value, note, type: "input"|"formula"|"section"|"blank"|"tbd")
ASSUMP = [
    (5,  "── TIER PRICING ──",                       None,   None,                                                                        "section"),
    (6,  "Creator monthly price",                    7.00,  "£7.00/mo — entry paid tier",                                                 "input"),
    (7,  "Studio monthly price",                    19.00,  "£19.00/mo — intended sweet spot",                                            "input"),
    (8,  "Enterprise monthly price",                45.00,  "£45.00/mo — team & power users",                                             "input"),
    (9,  None, None, None, "blank"),
    (10, "── TIER MIX (% of paying users) ──",       None,   None,                                                                        "section"),
    (11, "Creator share",                            0.20,  "20% of paying users — update from real data when available",                  "input"),
    (12, "Studio share",                             0.55,  "55% of paying users — intended sweet spot",                                   "input"),
    (13, "Enterprise share",                         0.25,  "25% of paying users",                                                         "input"),
    (14, "Blended ARPU (auto)",   "=$B$6*$B$11+$B$7*$B$12+$B$8*$B$13",  "Weighted avg revenue per paying user — updates automatically",  "formula"),
    (15, None, None, None, "blank"),
    (16, "── FIXED MONTHLY COSTS ──",                None,   None,                                                                        "section"),
    (17, "Vercel Pro (hosting)",                    16.00,  "Next.js hosting — Vercel Pro plan",                                           "input"),
    (18, "Supabase Pro (database)",                 20.00,  "Managed Postgres — required for commercial use",                              "input"),
    (19, "Domain & SSL (amortised)",                 1.25,  "~£15/year domain, amortised monthly",                                         "input"),
    (20, "Total fixed costs (auto)",   "=$B$17+$B$18+$B$19",  "Sum of all fixed costs — applies regardless of user count",               "formula"),
    (21, None, None, None, "blank"),
    (22, "── VARIABLE COSTS (per user / month) ──",  None,   None,                                                                        "section"),
    (23, "Explorer variable cost",                   0.005, "Storage & bandwidth for 3 trial videos — no AI, no screenshots",              "input"),
    (24, "Explorer AI cost",                         0.00,  "No AI on Explorer/trial tier",                                                "input"),
    (25, "Creator variable cost",                    0.01,  "Storage, bandwidth — no AI on Creator",                                       "input"),
    (26, "Creator AI cost",                          0.00,  "No AI features on Creator tier",                                              "input"),
    (27, "Studio variable cost",                     0.02,  "Storage + bandwidth",                                                         "input"),
    (28, "Studio AI cost",                           0.016, "40% adoption × 4 vids/mo × £0.01 Haiku per use",                             "input"),
    (29, "Enterprise variable cost",                 0.04,  "Higher storage & bandwidth",                                                  "input"),
    (30, "Enterprise AI cost",                       0.040, "50% adoption × 8 vids/mo × £0.01 Haiku per use",                             "input"),
    (31, None, None, None, "blank"),
    (32, "── PAYMENT PROCESSING (paid tiers only) ──", None, None,                                                                        "section"),
    (33, "Stripe percentage fee",                    0.022, "Stripe 1.5% + Clerk 0.7% overhead — Explorer pays nothing",                  "input"),
    (34, "Per-transaction fee (£)",                  0.20,  "£0.20 per monthly subscription — not charged on Explorer",                   "input"),
    (35, None, None, None, "blank"),
    (36, "── CLERK AUTH ──",                          None,  None,                                                                         "section"),
    (37, "Clerk monthly cost",                        0.00, "Free up to 10,000 MAU — revisit at scale",                                    "input"),
    (38, None, None, None, "blank"),
    (39, "── BLENDED MIX HELPERS (auto-calculated) ──", None, None,                                                                       "section"),
    (40, "Blended variable/user (auto)",  "=$B$25*$B$11+$B$27*$B$12+$B$29*$B$13",  "Weighted variable cost across paying tier mix",      "formula"),
    (41, "Blended AI cost/user (auto)",   "=$B$26*$B$11+$B$28*$B$12+$B$30*$B$13",  "Weighted AI cost across paying tier mix",            "formula"),
    (42, None, None, None, "blank"),
    (43, "── NOT YET MODELLED ──",                    None,  None,                                                                         "section"),
    (44, "Top-up credit revenue",                   "TBD", "£2.50 per +5 video pack — est. 20% Creator uptake — add when live",            "tbd"),
    (45, "Feature add-on revenue",                  "TBD", "Per-feature purchases on Creator — not yet priced",                            "tbd"),
    (46, "Annual plan discount",                    "TBD", "~15-20% discount for annual pre-pay — future option",                          "tbd"),
]

for row, label, value, note, ctype in ASSUMP:
    if ctype == "blank":
        continue
    if ctype == "section":
        ws_a.merge_cells(f"A{row}:C{row}")
        c = ws_a[f"A{row}"]
        c.value = label; c.fill = fill(C_HEAD2)
        c.font  = fnt(bold=True, color=C_WHITE, size=9); c.alignment = aln("left")
        ws_a.row_dimensions[row].height = 18
        continue

    bg = C_INPUT if ctype == "input" else ("EEEEFF" if ctype == "formula" else "FFF3E0")
    tc = C_DARK  if ctype != "tbd"   else C_MUTED

    c_a = ws_a.cell(row=row, column=1, value=label)
    c_a.fill = fill(bg); c_a.font = fnt(bold=(ctype == "formula"), color=tc, size=9)
    c_a.alignment = aln("left"); c_a.border = bdr()

    c_b = ws_a.cell(row=row, column=2, value=value)
    c_b.fill = fill(bg); c_b.font = fnt(bold=True, color=tc, size=9)
    c_b.alignment = aln("right"); c_b.border = bdr()
    # Number formats
    if "share" in (label or "").lower():                          c_b.number_format = PCT
    elif "stripe percentage" in (label or "").lower():            c_b.number_format = '0.00%'
    elif ctype in ("input", "formula") and isinstance(value, float): c_b.number_format = GBP
    elif ctype == "formula":                                       c_b.number_format = GBP

    c_c = ws_a.cell(row=row, column=3, value=note)
    c_c.fill = fill(bg); c_c.font = fnt(italic=True, color=C_MUTED, size=8)
    c_c.alignment = aln("left", wrap=True); c_c.border = bdr()

ws_a.merge_cells("A48:C48")
ws_a["A48"] = ("  Yellow cells = editable inputs you can change.   "
               "Blue/purple cells = auto-calculated (do not edit).   "
               "Orange cells = not yet modelled.")
ws_a["A48"].font = fnt(italic=True, color=C_MUTED, size=8)
ws_a["A48"].alignment = aln("left", wrap=True)
ws_a.row_dimensions[48].height = 20

ws_a.column_dimensions["A"].width = 36
ws_a.column_dimensions["B"].width = 16
ws_a.column_dimensions["C"].width = 58


# ══════════════════════════════════════════════════════════════════════════════
#  TIER SHEET BUILDER
# ══════════════════════════════════════════════════════════════════════════════
COL_HEADERS = [
    "Users", "Revenue (£)", "Fixed Costs (£)", "Stripe Fees (£)",
    "Variable (£)", "AI Costs (£)", "Total Costs (£)",
    "Net Profit (£)", "Margin (%)", "Annual Run Rate (£)", "Status"
]
COL_WIDTHS = [9, 14, 14, 13, 12, 12, 14, 14, 10, 17, 13]

def write_data_sheet(name, tab_color, subtitle, price_key, var_key, ai_key,
                     is_free=False):
    """
    Build a tier breakdown sheet.
    is_free=True: revenue fixed at 0, no Stripe fees (Explorer trial tier).
    """
    ws = wb.create_sheet(name)
    ws.sheet_properties.tabColor = tab_color
    ws.freeze_panes = "B4"

    ws.merge_cells("A1:K1")
    ws["A1"] = f"YT Transcriber — {name}"
    ws["A1"].fill = fill(C_DARK); ws["A1"].font = fnt(bold=True, color=C_WHITE, size=13)
    ws["A1"].alignment = aln("left"); ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:K2")
    ws["A2"] = f"{subtitle}     |     {TODAY}"
    ws["A2"].fill = fill(C_RED); ws["A2"].font = fnt(color=C_WHITE, size=9)
    ws["A2"].alignment = aln("left"); ws.row_dimensions[2].height = 16

    for ci, h in enumerate(COL_HEADERS, 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.fill = fill(C_HEAD2); c.font = fnt(bold=True, color=C_WHITE, size=9)
        c.alignment = aln("center", wrap=True); c.border = bdr()
    ws.row_dimensions[3].height = 28

    for ri, users in enumerate(USER_COUNTS, DATA_START):
        bg = C_LGREY if ri % 2 == 0 else "FFFFFF"

        ws.cell(row=ri, column=1, value=users).number_format = NUM

        # Col B — Revenue
        rev_f = "=0" if is_free else f"=A{ri}*{A(price_key)}"
        ws.cell(row=ri, column=2, value=rev_f).number_format = GBP

        # Col C — Fixed costs (always applies)
        ws.cell(row=ri, column=3, value=f"={A('fixed')}").number_format = GBP

        # Col D — Stripe fees (£0 for free tier — no payment processing)
        stripe_f = "=0" if is_free else f"=B{ri}*{A('stripe_pct')}+A{ri}*{A('stripe_fxd')}"
        ws.cell(row=ri, column=4, value=stripe_f).number_format = GBP

        # Col E — Variable cost
        ws.cell(row=ri, column=5, value=f"=A{ri}*{A(var_key)}").number_format = GBP

        # Col F — AI cost
        ws.cell(row=ri, column=6, value=f"=A{ri}*{A(ai_key)}").number_format = GBP

        # Col G — Total costs
        ws.cell(row=ri, column=7, value=f"=C{ri}+D{ri}+E{ri}+F{ri}").number_format = GBP

        # Col H — Net profit
        ws.cell(row=ri, column=8, value=f"=B{ri}-G{ri}").number_format = GBP

        # Col I — Margin (N/A for free tier but formula handles it)
        ws.cell(row=ri, column=9, value=f"=IF(B{ri}>0,H{ri}/B{ri},0)").number_format = PCT

        # Col J — Annual run rate
        ws.cell(row=ri, column=10, value=f"=H{ri}*12").number_format = GBP0

        # Col K — Status
        ws.cell(row=ri, column=11,
            value=f'=IF(H{ri}<0,"Loss",IF(H{ri}<20,"Break-even",IF(H{ri}<500,"Growing","Profitable")))'
        )

        for ci in range(1, 12):
            c = ws.cell(row=ri, column=ci)
            c.border = bdr()
            c.alignment = aln("center") if ci in (1, 11) else aln("right")
            if ci not in (8, 10, 11):
                c.fill = fill(bg); c.font = fnt(size=9)

    add_profit_cf(ws, "H", DATA_START, DATA_END)
    add_profit_cf(ws, "J", DATA_START, DATA_END)
    add_margin_cf(ws, "I", DATA_START, DATA_END)
    add_status_cf(ws, "K", DATA_START, DATA_END)

    for ci, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Footer note
    nr = DATA_END + 2
    ws.merge_cells(f"A{nr}:K{nr}")
    if is_free:
        note_text = (
            "Explorer is the free trial tier — Revenue and Stripe fees are always £0. "
            "Fixed infrastructure costs still apply. This section shows the net cost of running free/trial users. "
            "Variable cost per user reflects storage & bandwidth for up to 3 trial video transcriptions."
        )
    else:
        note_text = (
            "All values driven by Assumptions sheet. Edit yellow input cells there and this sheet updates automatically. "
            "Fixed costs are shared infrastructure costs that apply regardless of how many users are on this tier."
        )
    ws[f"A{nr}"] = note_text
    ws[f"A{nr}"].font = fnt(italic=True, color=C_MUTED, size=8)
    ws[f"A{nr}"].alignment = aln("left", wrap=True)
    ws.row_dimensions[nr].height = 32

    return ws


# ── Build individual tier sheets ──────────────────────────────────────────────
write_data_sheet(
    "Explorer Tier", "9E9E9E",
    "Free trial — 3 videos total, no revenue, no Stripe fees — shows cost of running free users",
    "explorer_var", "explorer_var", "explorer_ai",   # price_key unused (is_free=True)
    is_free=True
)
write_data_sheet(
    "Creator Tier", "4FC3F7",
    "Entry paid tier — 15 videos/mo, core transcript features, no AI, no screenshots in ZIP",
    "creator_price", "creator_var", "creator_ai"
)
write_data_sheet(
    "Studio Tier", "CE93D8",
    "Sweet spot — 60 videos/mo, all features, AI chapters included",
    "studio_price", "studio_var", "studio_ai"
)
write_data_sheet(
    "Enterprise Tier", "FFCC02",
    "Power tier — Unlimited videos, team workspace, priority processing",
    "ent_price", "ent_var", "ent_ai"
)

# ── Blended Mix sheet ─────────────────────────────────────────────────────────
ws_blend = write_data_sheet(
    "Blended Mix", C_RED,
    "Paying users only — 20% Creator / 55% Studio / 25% Enterprise — adjust mix in Assumptions",
    "arpu", "blend_var", "blend_ai"
)

# Bar chart on Blended Mix
chart = BarChart()
chart.type   = "col"
chart.title  = "Blended: Revenue vs Total Costs vs Net Profit"
chart.y_axis.title = "Amount (£)"; chart.x_axis.title = "Paying Users"
chart.style  = 10; chart.width = 22; chart.height = 14
ref_rev    = Reference(ws_blend, min_col=2, min_row=3, max_row=DATA_END)
ref_cost   = Reference(ws_blend, min_col=7, min_row=3, max_row=DATA_END)
ref_profit = Reference(ws_blend, min_col=8, min_row=3, max_row=DATA_END)
ref_users  = Reference(ws_blend, min_col=1, min_row=DATA_START, max_row=DATA_END)
chart.add_data(ref_rev,    titles_from_data=True)
chart.add_data(ref_cost,   titles_from_data=True)
chart.add_data(ref_profit, titles_from_data=True)
chart.set_categories(ref_users)
chart.series[0].graphicalProperties.solidFill = "3B82F6"
chart.series[1].graphicalProperties.solidFill = "E53935"
chart.series[2].graphicalProperties.solidFill = "4CAF50"
ws_blend.add_chart(chart, f"A{DATA_END + 4}")


# ══════════════════════════════════════════════════════════════════════════════
#  SUMMARY SHEET  (inserted first, tab 0)
# ══════════════════════════════════════════════════════════════════════════════
ws_s = wb.create_sheet("Summary", 0)
ws_s.sheet_properties.tabColor = C_DARK
ws_s.freeze_panes = "B5"

ws_s.merge_cells("A1:G1")
ws_s["A1"] = "YT Transcriber — Financial Model"
ws_s["A1"].fill = fill(C_DARK); ws_s["A1"].font = fnt(bold=True, color=C_WHITE, size=16)
ws_s["A1"].alignment = aln("left"); ws_s.row_dimensions[1].height = 36
ws_s.merge_cells("A2:G2")
ws_s["A2"] = f"Edit Assumptions sheet (yellow cells) to update all figures automatically     |     {TODAY}"
ws_s["A2"].fill = fill(C_RED); ws_s["A2"].font = fnt(color=C_WHITE, size=9)
ws_s["A2"].alignment = aln("left"); ws_s.row_dimensions[2].height = 16

SUM_HDRS = ["Users", "Monthly Revenue", "Total Costs", "Net Profit", "Margin", "Annual Profit", "Status"]

def write_summary_section(ws, start_row, section_label, price_key, var_key, ai_key,
                           is_free=False):
    ws.merge_cells(f"A{start_row}:G{start_row}")
    ws[f"A{start_row}"] = section_label
    ws[f"A{start_row}"].fill = fill(C_HEAD2)
    ws[f"A{start_row}"].font = fnt(bold=True, color=C_WHITE, size=9)
    ws[f"A{start_row}"].alignment = aln("center")
    ws.row_dimensions[start_row].height = 20
    r = start_row + 1

    for ci, h in enumerate(SUM_HDRS, 1):
        c = ws.cell(row=r, column=ci, value=h)
        c.fill = fill("3A3A3A"); c.font = fnt(bold=True, color=C_WHITE, size=9)
        c.alignment = aln("center", wrap=True); c.border = bdr()
    ws.row_dimensions[r].height = 22
    r += 1

    for users in USER_COUNTS:
        bg = C_LGREY if r % 2 == 0 else "FFFFFF"
        ws.cell(row=r, column=1, value=users).number_format = NUM

        # Revenue
        rev_f = "=0" if is_free else f"=A{r}*{A(price_key)}"
        ws.cell(row=r, column=2, value=rev_f).number_format = GBP

        # Total costs (all-in-one)
        stripe_part = "" if is_free else f"+B{r}*{A('stripe_pct')}+A{r}*{A('stripe_fxd')}"
        ws.cell(row=r, column=3,
            value=f"={A('fixed')}{stripe_part}+A{r}*{A(var_key)}+A{r}*{A(ai_key)}"
        ).number_format = GBP

        # Net profit / Margin / Annual / Status
        ws.cell(row=r, column=4, value=f"=B{r}-C{r}").number_format = GBP
        ws.cell(row=r, column=5, value=f"=IF(B{r}>0,D{r}/B{r},0)").number_format = PCT
        ws.cell(row=r, column=6, value=f"=D{r}*12").number_format = GBP0
        ws.cell(row=r, column=7,
            value=f'=IF(D{r}<0,"Loss",IF(D{r}<20,"Break-even",IF(D{r}<500,"Growing","Profitable")))'
        )

        for ci in range(1, 8):
            c = ws.cell(row=r, column=ci)
            c.border = bdr()
            c.alignment = aln("center") if ci in (1, 7) else aln("right")
            if ci not in (4, 6, 7):
                c.fill = fill(bg); c.font = fnt(size=9)
        r += 1

    add_profit_cf(ws, "D", start_row + 2, r - 1)
    add_profit_cf(ws, "F", start_row + 2, r - 1)
    add_margin_cf(ws, "E", start_row + 2, r - 1)
    add_status_cf(ws, "G", start_row + 2, r - 1)

    return r + 1

sr = 4
sr = write_summary_section(ws_s, sr,
    "EXPLORER TIER  —  Free trial (3 videos)  |  Revenue: £0  |  Shows cost of running free/trial users",
    "explorer_var", "explorer_var", "explorer_ai", is_free=True)
sr = write_summary_section(ws_s, sr,
    "BLENDED MIX  —  Paying users: 20% Creator / 55% Studio / 25% Enterprise  (adjust in Assumptions)",
    "arpu", "blend_var", "blend_ai")
sr = write_summary_section(ws_s, sr,
    "CREATOR TIER  —  £7/mo per user",
    "creator_price", "creator_var", "creator_ai")
sr = write_summary_section(ws_s, sr,
    "STUDIO TIER  —  £19/mo per user",
    "studio_price", "studio_var", "studio_ai")
sr = write_summary_section(ws_s, sr,
    "ENTERPRISE TIER  —  £45/mo per user",
    "ent_price", "ent_var", "ent_ai")

ws_s.merge_cells(f"A{sr}:G{sr}")
ws_s[f"A{sr}"] = (
    "All figures driven by Assumptions sheet — edit yellow input cells there to update this entire workbook.  "
    "Explorer shows infrastructure cost of free/trial users — Revenue and Stripe fees are always £0.  "
    "Top-up credit revenue not yet included."
)
ws_s[f"A{sr}"].font = fnt(italic=True, color=C_MUTED, size=8)
ws_s[f"A{sr}"].alignment = aln("left", wrap=True)
ws_s.row_dimensions[sr].height = 32

for ci, w in enumerate([14, 16, 16, 14, 10, 16, 13], 1):
    ws_s.column_dimensions[get_column_letter(ci)].width = w

wb.save(OUTPUT)
print(f"Excel saved to {OUTPUT}")
